import openpyxl

class ExcelReader:

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def read_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return data

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def close(self):
        self.workbook.close()

# Test code
# Example usage:
excel_file_path = "path/to/your/excel/file.xlsx"
excel_reader = ExcelReader(excel_file_path)

sheet_names = excel_reader.get_sheet_names()
print("Sheet names:", sheet_names)

sheet_name = sheet_names[0]  # Choose the sheet you want to read
sheet_data = excel_reader.read_sheet(sheet_name)
print(f"Data from '{sheet_name}' sheet:")
for row in sheet_data:
    print(row)

excel_reader.close()